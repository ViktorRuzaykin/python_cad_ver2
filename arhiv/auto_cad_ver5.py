from pyautocad import Autocad, APoint
import time
import openpyxl
import random
from tkinter.filedialog import *
from tkinter import messagebox as mb
from tkinter import ttk
from tkinter.ttk import Progressbar, Combobox, Checkbutton


def autocad_len():

    try:
        acad1 = Autocad()
        mSp_1 = acad1.doc
    except:
        mb.showerror(
            "AutoCAD не доступен",
            f'AutoCAD не доступен. Пробуй открыть AutoCAD, перезапустить его или создать чертеж.')

    h_step = 50  # шаг строк подвала(линии)

    f_pk_int = []
    f_pk_float = []
    f_mark_ditch = []
    f_mark_ditch_fact = []
    f_mark_pillow = []
    f_mark_pillow_fact = []
    f_mark_pipe = []
    f_mark_pipe_fact = []
    f_mark_filling = []
    f_mark_filling_fact = []
    f_mark_earth = []

    def calc_file():
        """
        Функция для расчета проектных и  фактических отметок по траншеии, подсыпке, трубе, обсыпке. За основу берется таблица с 3-мя
        столбцами: пк, плюс, Нтрубы, Нземли
        """
        # очищаем списки с отметками, на случай если за сесею выполняем расчет отметок для тругого файла
        f_pk_int.clear()
        f_pk_float.clear()
        f_mark_ditch.clear()
        f_mark_ditch_fact.clear()
        f_mark_pillow.clear()
        f_mark_pillow_fact.clear()
        f_mark_pipe.clear()
        f_mark_pipe_fact.clear()
        f_mark_filling.clear()
        f_mark_filling_fact.clear()
        f_mark_earth.clear()
        open_file_1 = askopenfilename()  # переменная выбранного файла для расчета отметок
        # обработка CSV файла
        wb = openpyxl.reader.excel.load_workbook(filename=open_file_1)
        sheet = wb.active
        nb_row = sheet.max_row  # число строк в файле

        count = 1
        for num in range(nb_row + 1):
            if sheet[f'A{count}'].value == None:
                pass
            else:
                f_pk_int.append(sheet[f'A{count}'].value)
                f_pk_float.append(sheet[f'B{count}'].value)
                f_mark_pipe.append(sheet[f'C{count}'].value)
                f_mark_earth.append(sheet[f'D{count}'].value)
                f_mark_ditch.append(float(sheet[f'C{count}'].value) - 1.62)
                f_mark_pillow.append(float(sheet[f'C{count}'].value) - 1.42)
                f_mark_filling.append(float(sheet[f'C{count}'].value) + 0.20)
            count += 1

        for i in range(len(f_mark_pipe)):
            f_mark_ditch_fact.append(f_mark_ditch[i] - (random.randint(1, 9) / 100))
            f_mark_filling_fact.append(f_mark_filling[i] + (random.randint(1, 4) / 100))

            f_pillow_fact = f_mark_pillow[i] - (random.randint(1, 6) / 100)
            dx_pillow_fact_ditch = f_pillow_fact - f_mark_ditch_fact[i]
            if dx_pillow_fact_ditch < 0.20:
                f_mark_pillow_fact.append(f_pillow_fact + (0.20 - dx_pillow_fact_ditch))
            else:
                f_mark_pillow_fact.append(f_pillow_fact)
            f_mark_pipe_fact.append(f_mark_pillow_fact[i] + 1.42)

        new_file_calc = f'{int(f_pk_int[0])}+{int(f_pk_float[0])}-{int(f_pk_int[-1])}+{int(f_pk_float[-1])}_calc.xlsx'
        wb_write = openpyxl.Workbook(new_file_calc)
        wb_write.save(new_file_calc)
        wb_write = openpyxl.reader.excel.load_workbook(filename=new_file_calc)
        sheet_write = wb_write.active

        count_v = 2
        for value in range(len(f_pk_int)):
            pk_int = sheet_write[f'A{count_v}']
            pk_int.value = f_pk_int[value]
            pk_float = sheet_write[f'B{count_v}']
            pk_float.value = f'{f_pk_float[value]}'
            mark_pipe = sheet_write[f'C{count_v}']
            mark_pipe.value = f'{f_mark_pipe[value]}'
            mark_earth = sheet_write[f'D{count_v}']
            mark_earth.value = f'{f_mark_earth[value]}'
            mark_ditch = sheet_write[f'E{count_v}']
            mark_ditch.value = f'{f_mark_ditch[value]}'
            mark_ditch_fact = sheet_write[f'F{count_v}']
            mark_ditch_fact.value = f'{f_mark_ditch_fact[value]}'
            mark_pillow = sheet_write[f'G{count_v}']
            mark_pillow.value = f'{f_mark_pillow[value]}'
            mark_pillow_fact = sheet_write[f'H{count_v}']
            mark_pillow_fact.value = f'{f_mark_pillow_fact[value]}'
            mark_pipe_fact = sheet_write[f'I{count_v}']
            mark_pipe_fact.value = f'{f_mark_pipe_fact[value]}'
            mark_filling = sheet_write[f'J{count_v}']
            mark_filling.value = f'{f_mark_filling[value]}'
            mark_filling_fact = sheet_write[f'K{count_v}']
            mark_filling_fact.value = f'{f_mark_filling_fact[value]}'
            count_v += 1

        tuple_column = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K')
        name_column = ['пк', 'плюс', 'H трубы', 'Н земли', 'Н транш', 'Н ф.транш', 'Н подуш', 'Н ф.подуш', 'Н ф.трубы', 'Н обсып', 'Н ф.обсып']
        for count in range(len(tuple_column)):
            name_column_w = sheet_write[f'{tuple_column[count]}1']
            name_column_w.value = f'{name_column[count]}'

        wb_write.save(new_file_calc)

        name_file = f'           {f_pk_int[0]}+{f_pk_float[0]}0 - {f_pk_int[-1]}+{f_pk_float[-1]}0          '
        lbl_5.configure(text=name_file, background="#47fa41")
        mb.showinfo(
            "Внимание!",
            f'Добавлен файл {new_file_calc}.')

    pk_int = []
    pk_float = []
    mark_ditch = []
    mark_ditch_fact = []
    mark_pillow = []
    mark_pillow_fact = []
    mark_pipe = []
    mark_pipe_fact = []
    mark_filling = []
    mark_filling_fact = []
    mark_earth = []

    def file():
        pk_int.clear()
        pk_float.clear()
        mark_ditch.clear()
        mark_ditch_fact.clear()
        mark_pillow.clear()
        mark_pillow_fact.clear()
        mark_pipe.clear()
        mark_pipe_fact.clear()
        mark_filling.clear()
        mark_filling_fact.clear()
        mark_earth.clear()
        open_file = askopenfilename()  # переменная выбранного файла c отметками факт/проект
        # обработка xlsx файла
        wb = openpyxl.reader.excel.load_workbook(filename=open_file)
        sheet = wb.active
        nb_row = sheet.max_row  # число строк в файле

        count = 2
        for num in range(nb_row + 1):
            if sheet[f'A{count}'].value == None:
                pass
            else:
                pk_int.append(str(sheet[f'A{count}'].value))
                pk_float.append(str(sheet[f'B{count}'].value))
                mark_pipe.append(float(sheet[f'C{count}'].value))
                mark_earth.append(float(sheet[f'D{count}'].value))
                mark_ditch.append(float(sheet[f'E{count}'].value))
                mark_ditch_fact.append(float(sheet[f'F{count}'].value))
                mark_pillow.append(float(sheet[f'G{count}'].value))
                mark_pillow_fact.append(float(sheet[f'H{count}'].value))
                mark_pipe_fact.append(float(sheet[f'I{count}'].value))
                mark_filling.append(float(sheet[f'J{count}'].value))
                mark_filling_fact.append(float(sheet[f'K{count}'].value))
            count += 1

        # редактируем список pk_float, заменяем '0' на '00'
        for i in range(len(pk_float)):
            if pk_float[i] == '0':
                pk_float[i] = '00'

        name_file = f'           {pk_int[0]}+{pk_float[0]}0 - {pk_int[-1]}+{pk_float[-1]}         '
        lbl_1.configure(text=name_file, background="#47fa41")

    def calc():
        def style_text():
            for text in acad.iter_objects('Text'):
                text.StyleName = combo.get()
        try:
            acad = Autocad()
            mSp = acad.model
        except OSError:
            mb.showerror(
                "AutoCAD не доступен",
                f'Пробуй открыть AutoCAD или перезапустить его.')

        bar['value'] = 0
        window.update_idletasks()

        # переводим ПК в метры
        distance_pk = []
        for pk in range(len(pk_float)):
            if 0 < float(pk_float[pk]) < 10:
                distance_pk.append(pk_int[pk] + '0' + pk_float[pk])
            else:
                distance_pk.append(pk_int[pk] + pk_float[pk])
        # текст для шапки профиля трубы
        text_basement_pipe = ['Отметки земли, м', 'Проектные отметки', 'верха трубы, м', 'Фактические отметки',
                              'верха трубы, м', 'Разница в отметках, м', 'Откос',
                              'Ширина дна траншеи', 'Тип прокладки', 'Длина', 'Уклон', 'Расстояние, м', 'Пикеты, м']

        # текст для шапки профиля траншеи
        text_basement_transh = ['Отметки земли, м', 'Проектные отметки', 'дна траншеи, м', 'Фактические отметки',
                                'дна траншеи, м', 'Разница в отметках, м', 'Глубина траншеи, м', 'Откос',
                                'Ширина дна траншеи', 'Тип прокладки', 'Расстояние, м', 'Пикеты, м']

        # текст для шапки профиля подушки
        text_basement_pillow = ['Фактические отметки', 'дна траншеи, м', 'Проектные отметки', 'верха подсыпки, м',
                                'Фактические отметки', 'верха подсыпки, м', 'Толщина слоя подсыпки, м', 'Откос',
                                'Ширина дна траншеи', 'Тип прокладки', 'Расстояние, м', 'Пикеты, м']

        # текст для шапки профиля обсыпки
        text_basement_filling = ['Проектные отметки', 'верха обсыпки, м', 'Фактические отметки', 'верха обсыпки, м',
                                 'Фактические отметки', 'верха трубы, м', 'Толщина слоя', 'над трубой, м', 'Откос',
                                 'Ширина дна траншеи', 'Тип прокладки', 'Расстояние, м', 'Пикеты, м']

        # контур шапки подвала
        def line_basement(height, width, dx):
            mSp.AddLine(APoint(0 + dx, 0), APoint(0 + dx, height))
            mSp.AddLine(APoint(0 + dx, height), APoint(width + dx, height))
            mSp.AddLine(APoint(width + dx, height), APoint(width + dx, 0))
            mSp.AddLine(APoint(0 + dx, 0), APoint(width + dx, 0))

        # горизонтальные черты в шапке подвала
        def line_basement_gorizont(h_step, number_line, dx):
            k = 0
            for line in range(0, number_line):
                th = k + h_step
                mSp.AddLine(APoint(0 + dx, th), APoint(264 + dx, th))
                k += h_step

        # подпись строк в шапке подвала
        def text_basement_hat(text, ditch, dx):
            for i in range(len(ditch)):
                p = APoint(float(ditch[i][0]), float(ditch[i][1]))
                mSp.AddText(text[i], p, 14)

        def line_longitudinal(s, dx_ditch):
            g = 0
            for e in range(0, s):
                th_2 = g + h_step
                mSp.AddLine(APoint(264 + dx_ditch, th_2), APoint(profile_distance_last + 264 + dx_ditch, th_2))
                g += h_step
            mSp.AddLine(APoint(264 + dx_ditch, 0), APoint(264 + profile_distance_last + dx_ditch, 0))

        # интерполяция начала и конца профиля
        data_1 = str(input_start_pk.get())
        data_2 = str(input_end_pk.get())

        def star_end_point():
            dict_start_end = {}
            try:
                #data_1 = '941+00'
                input_start_pk = data_1.split('+')
                start_pk = ''.join(input_start_pk)
                dict_start_end['start_pk'] = start_pk

                #data_2 = '945+74'#
                input_end_pk = data_2.split('+')

                end_pk = ''.join(input_end_pk)
                dict_start_end['end_pk'] = end_pk

                dict_start_end['data_1'] = data_1
                dict_start_end['data_2'] = data_2

                dict_start_end['first_pk_int'] = input_start_pk[0]
                dict_start_end['first_pk_float'] = input_start_pk[1]
                dict_start_end['end_pk_int'] = input_end_pk[0]
                dict_start_end['end_pk_float'] = input_end_pk[1]
            except IndexError:
                mb.showerror(
                    'Ошибка',
                    f'Не верный ПК/Введите ПК')
            except KeyError:
                mb.showerror(
                    'Ошибка',
                    f'Не верный ПК/Введите ПК')

            for pk_st in range(len(distance_pk)):
                if start_pk == distance_pk[pk_st]:
                    dict_start_end['int_st'] = True
                    dict_start_end['p_start'] = pk_st
                    #print(pk_st)

                if end_pk == distance_pk[pk_st]:
                    dict_start_end['int_end'] = True
                    dict_start_end['p_end'] = pk_st+1
                    #print(pk_st)

                if distance_pk[pk_st - 1] < start_pk < distance_pk[pk_st]:
                    dict_start_end['int_st'] = False
                    dict_start_end['p_start'] = pk_st
                    distance_1 = round(float(distance_pk[pk_st]) - float(start_pk), 2)
                    dict_start_end['distance_1'] = distance_1
                    first_earth = abs(round(((((mark_earth[pk_st] - mark_earth[pk_st - 1])) / (float(distance_pk[pk_st]) - float(distance_pk[pk_st - 1]))) * distance_1) - mark_earth[pk_st], 2))
                    dict_start_end['first_earth'] = first_earth
                    first_pipe = abs(round(((((mark_pipe[pk_st] - mark_pipe[pk_st - 1])) / (float(distance_pk[pk_st]) - float(distance_pk[pk_st - 1]))) * distance_1) - mark_pipe[pk_st], 2))
                    dict_start_end['first_pipe'] = first_pipe
                    first_pillow = round(first_pipe - 1.42, 2)
                    dict_start_end['first_pillow'] = first_pillow
                    first_filling = round(first_pipe + 0.20, 2)
                    dict_start_end['first_filling'] = first_filling
                    first_ditch = round(first_pipe - 1.62, 2)
                    dict_start_end['first_ditch'] = first_ditch
                    first_pipe_fact = round(first_pipe - 0.03, 2)
                    dict_start_end['first_pipe_fact'] = first_pipe_fact
                    first_pillow_fact = round(first_pillow - 0.03, 2)
                    dict_start_end['first_pillow_fact'] = first_pillow_fact
                    first_ditch_fact = round(first_ditch - 0.03, 2)
                    dict_start_end['first_ditch_fact'] = first_ditch_fact
                    first_filling_fact = round(first_filling + 0.03, 2)
                    dict_start_end['first_filling_fact'] = first_filling_fact

                if distance_pk[pk_st - 1] < end_pk < distance_pk[pk_st]:
                    dict_start_end['int_end'] = False
                    dict_start_end['p_end'] = pk_st
                    distance_2 = round(float(end_pk) - float(distance_pk[pk_st - 1]), 2)
                    dict_start_end['distance_2'] = distance_2
                    end_earth = abs(round(((((mark_earth[pk_st] - mark_earth[pk_st - 1])) / (float(distance_pk[pk_st]) - float(distance_pk[pk_st - 1]))) * distance_2) + mark_earth[pk_st-1], 2))
                    dict_start_end['end_earth'] = end_earth
                    end_pipe = abs(round(((((mark_pipe[pk_st] - mark_pipe[pk_st - 1])) / (float(distance_pk[pk_st]) - float(distance_pk[pk_st - 1]))) * distance_2) + mark_pipe[pk_st-1], 2))
                    dict_start_end['end_pipe'] = end_pipe
                    end_pillow = round(end_pipe - 1.42, 2)
                    dict_start_end['end_pillow'] = end_pillow
                    end_filling = round(end_pipe + 0.20, 2)
                    dict_start_end['end_filling'] = end_filling
                    end_ditch = round(end_pipe - 1.62, 2)
                    dict_start_end['end_ditch'] = end_ditch
                    end_pipe_fact = round(end_pipe - 0.03, 2)
                    dict_start_end['end_pipe_fact'] = end_pipe_fact
                    end_pillow_fact = round(end_pillow - 0.03, 2)
                    dict_start_end['end_pillow_fact'] = end_pillow_fact
                    end_ditch_fact = round(end_ditch - 0.03, 2)
                    dict_start_end['end_ditch_fact'] = end_ditch_fact
                    end_filling_fact = round(end_filling + 0.03, 2)
                    dict_start_end['end_filling_fact'] = end_filling_fact
            return dict_start_end

        s_e_p = star_end_point()

        p_start = s_e_p['p_start']
        p_end = s_e_p['p_end']
        if s_e_p['int_st'] == False:
            distance_1 = s_e_p['distance_1']
        else:
            distance_1 = 0
        if s_e_p['int_end'] == False:
            distance_2 = s_e_p['distance_2']
        else:
            distance_2 = 0

        new_distance_pk = distance_pk[p_start:p_end]
        if s_e_p['int_st'] == False:
            new_distance_pk.insert(0, s_e_p['start_pk'])
        if s_e_p['int_end'] == False:
            new_distance_pk.append(s_e_p['end_pk'])

        new_mark_earth = mark_earth[p_start:p_end]
        if s_e_p['int_st'] == False:
            new_mark_earth.insert(0, s_e_p['first_earth'])
        if s_e_p['int_end'] == False:
            new_mark_earth.append(s_e_p['end_earth'])

        new_mark_ditch = mark_ditch[p_start:p_end]
        if s_e_p['int_st'] == False:
            new_mark_ditch.insert(0, s_e_p['first_ditch'])
        if s_e_p['int_end'] == False:
            new_mark_ditch.append(s_e_p['end_ditch'])

        new_mark_ditch_fact = mark_ditch_fact[p_start:p_end]
        if s_e_p['int_st'] == False:
            new_mark_ditch_fact.insert(0, s_e_p['first_ditch_fact'])
        if s_e_p['int_end'] == False:
            new_mark_ditch_fact.append(s_e_p['end_ditch_fact'])

        new_mark_pillow = mark_pillow[p_start:p_end]
        if s_e_p['int_st'] == False:
            new_mark_pillow.insert(0, s_e_p['first_pillow'])
        if s_e_p['int_end'] == False:
            new_mark_pillow.append(s_e_p['end_pillow'])

        new_mark_pillow_fact = mark_pillow_fact[p_start:p_end]
        if s_e_p['int_st'] == False:
            new_mark_pillow_fact.insert(0, s_e_p['first_pillow_fact'])
        if s_e_p['int_end'] == False:
            new_mark_pillow_fact.append(s_e_p['end_pillow_fact'])

        new_mark_pipe = mark_pipe[p_start:p_end]
        if s_e_p['int_st'] == False:
            new_mark_pipe.insert(0, s_e_p['first_pipe'])
        if s_e_p['int_end'] == False:
            new_mark_pipe.append(s_e_p['end_pipe'])

        new_mark_pipe_fact = mark_pipe_fact[p_start:p_end]
        if s_e_p['int_st'] == False:
            new_mark_pipe_fact.insert(0, s_e_p['first_pipe_fact'])
        if s_e_p['int_end'] == False:
            new_mark_pipe_fact.append(s_e_p['end_pipe_fact'])

        new_mark_filling = mark_filling[p_start:p_end]
        if s_e_p['int_st'] == False:
            new_mark_filling.insert(0, s_e_p['first_filling'])
        if s_e_p['int_end'] == False:
            new_mark_filling.append(s_e_p['end_filling'])

        new_mark_filling_fact = mark_filling_fact[p_start:p_end]
        if s_e_p['int_st'] == False:
            new_mark_filling_fact.insert(0, s_e_p['first_filling_fact'])
        if s_e_p['int_end'] == False:
            new_mark_filling_fact.append(s_e_p['end_filling_fact'])

        new_pk_int = pk_int[p_start:p_end]
        if s_e_p['int_st'] == False:
            new_pk_int.insert(0, s_e_p['first_pk_int'])
        if s_e_p['int_end'] == False:
            new_pk_int.append(s_e_p['end_pk_int'])

        new_pk_float = pk_float[p_start:p_end]
        if s_e_p['int_st'] == False:
            new_pk_float.insert(0, s_e_p['first_pk_float'])
        if s_e_p['int_end'] == False:
            new_pk_float.append(s_e_p['end_pk_float'])

        # расчет длинны профиля (нач. ПК - конеч. ПК)
        a = float(new_pk_int[0] + new_pk_float[0])
        b = float(new_pk_int[-1] + new_pk_float[-1])
        profile_distance = b - a
        profile_distance_last = float(profile_distance + 40)

        # расчет смещения для профилей
        dx_pipe = profile_distance_last + 500
        dx_ditch = dx_pipe
        dx_pillow = dx_pipe + profile_distance_last + 500
        dx_filling = dx_pillow + profile_distance_last + 500
        dx_ditch_earth = (new_mark_earth[0] - new_mark_ditch[0]) * 25

        coordinates_text_basement_pipe = [[83.75 + dx_pipe, 420.16], [61.21 + dx_pipe, 381.63],
                                          [109.57 + dx_pipe, 362.79], [40.01 + dx_pipe, 330.54],
                                          [108.31 + dx_pipe, 309.99], [33.89 + dx_pipe, 270.16],
                                          [94.81 + dx_pipe, 231.78], [47.29 + dx_pipe, 210.97],
                                          [112.15 + dx_pipe, 168.63], [6.45 + dx_pipe, 110.89],
                                          [188.99 + dx_pipe, 128.31], [116.18 + dx_pipe, 68.49], [153.10 + dx_pipe, 21.65]]

        coordinates_text_basement_pillow = [[39.41 + dx_pillow, 378.69], [105.21 + dx_pillow, 359.85],
                                            [61.294 + dx_pillow, 330.54], [81.91 + dx_pillow, 309.99],
                                            [39.66 + dx_pillow, 280.54], [81.72 + dx_pillow, 259.99],
                                            [12.02 + dx_pillow, 219.802], [94.810 + dx_pillow, 181.780],
                                            [47.290 + dx_pillow, 160.970], [111.82 + dx_pillow, 118.630],
                                            [116.180 + dx_pillow, 69.180], [153.100 + dx_pillow, 18.740]]

        coordinates_text_basement_filling = [[61.42 + dx_filling, 378.99], [91.78 + dx_filling, 360.15],
                                             [39.72 + dx_filling, 330.54], [92.11 + dx_filling, 309.99],
                                             [39.72 + dx_filling, 280.54], [107.83 + dx_filling, 259.99],
                                             [119.79 + dx_filling, 230.54], [118.88 + dx_filling, 209.99],
                                             [94.81 + dx_filling, 181.78], [50.32 + dx_filling, 160.97],
                                             [112.16 + dx_filling, 118.63], [116.18 + dx_filling, 69.18],
                                             [153.10 + dx_filling, 18.7]]

        coordinates_text_basement_transh = [[83.66, 418.58], [61.21, 381.63],
                                            [104.80, 362.79], [39.11, 330.54],
                                            [104.79, 309.99], [33.89, 270.16],
                                            [62.96, 221.11], [94.81, 181.78],
                                            [47.29, 160.97], [111.77, 118.63],
                                            [116.18, 69.18], [153.10, 18.74]]

        def basement_all_ditch():
            # начальня риска в строке "Растояния"
            mSp.AddLine(APoint(304.00, 50.00), APoint(304.00, 70.00))
            # начальная отметка земли
            mSp.AddText(new_mark_earth[0], APoint(304.00, 404.75), 10).rotation = 1.57
            # начальная проектная отметка траншеии
            mSp.AddText(format(new_mark_ditch[0], '.2f'), APoint(304.00, 354.33), 10).rotation = 1.57
            # начальная фактическая отметка траншеии
            mSp.AddText(format(round(new_mark_ditch_fact[0], 2), '.2f'), APoint(304.00, 304.35), 10).rotation = 1.57
            # начальное отклонение
            mSp.AddText(round(float(new_mark_ditch_fact[0])-float(new_mark_ditch[0]), 2), APoint(304.00, 256.83), 10).rotation = 1.57
            # начальная глубина разработки
            mSp.AddText(format(round(float(new_mark_earth[0]) - float(new_mark_ditch_fact[0]), 2), '.2f'), APoint(304.00, 209.99), 10).rotation = 1.57
            # начальный ПК в строке "Пикеты"
            pk_start = new_pk_int[0] + '+' + new_pk_float[0]
            mSp.AddText(pk_start, APoint(280.00, 32), 10)
            p = 304.00
            for i in range(len(new_distance_pk) - 1):
                pk_dist = round(float(new_distance_pk[i + 1]) - float(new_distance_pk[i]), 2)
                dh_x = p + pk_dist
                mSp.AddLine(APoint(dh_x, 70.00), APoint(dh_x, 50.00))
                # подпись расстояния
                otstup_text = pk_dist - (pk_dist / 2 + 5)
                mSp.AddText(format(pk_dist, '.2f'), APoint(dh_x - otstup_text, 61.30), 10).rotation = 1.57
                p += pk_dist
                # отметки верха земли
                mSp.AddText(format(new_mark_earth[i + 1], '.2f'), APoint(dh_x, 404.75), 10).rotation = 1.57
                # проектные отметки дна траншеии
                mSp.AddText(format(float(new_mark_ditch[i + 1]), '.2f'), APoint(dh_x, 354.33), 10).rotation = 1.57
                # фактические отметки дна траншеи
                mSp.AddText(format(new_mark_ditch_fact[i + 1], '.2f'), APoint(dh_x, 304.35), 10).rotation = 1.57
                # отклонения от проекта по дну трашеи
                mSp.AddText(round(new_mark_ditch_fact[i + 1] - new_mark_ditch[i + 1], 2), APoint(dh_x, 256.83), 10).rotation = 1.57
                # глубина разработки траншеи
                development_depth = format(round(new_mark_earth[i + 1] - new_mark_ditch_fact[i + 1], 2), '.2f')
                mSp.AddText(development_depth, APoint(dh_x, 209.99), 10).rotation = 1.57

                if len(new_distance_pk)-2 == i:
                    l_prof = float(new_distance_pk[-1]) - float(new_distance_pk[0]) - 15
                    pk_end = new_pk_int[-1] + '+' + new_pk_float[-1]
                    mSp.AddText(pk_end, APoint(304.00+l_prof, 32), 10)
                else:
                    piket = new_pk_int[i + 1]
                    cutoff = len(piket) - 1
                    if new_pk_float[i + 1] == '00':
                        mSp.AddLine(APoint(dh_x, 100.00), APoint(dh_x, 50.00))
                        if piket[cutoff:] == '0' or piket[cutoff:] == '5':
                            mSp.AddText(piket, APoint(dh_x - 15, 32), 10)
                        else:
                            mSp.AddText(piket[cutoff:], APoint(dh_x - 3.24, 32), 10)
            if new_pk_float[0] == '00':
                mSp.AddLine(APoint(304, 100.00), APoint(304, 50.00))

        def basement_all_pillow():
            # начальня риска в строке "Растояния"
            mSp.AddLine(APoint(304.00 + dx_pillow, 50.00), APoint(304.00 + dx_pillow, 70.00))
            # начальная фактическая отметка дна траншеии
            mSp.AddText(format(new_mark_ditch_fact[0], '.2f'), APoint(304.00 + dx_pillow, 354.33), 10).rotation = 1.57
            # начальная проектная отметка верха подушки
            mSp.AddText(format(new_mark_pillow[0], '.2f'), APoint(304.00 + dx_pillow, 304.35), 10).rotation = 1.57
            # начальная фактическая отметка верха подушки
            mSp.AddText(format(round(new_mark_pillow_fact[0], 2), '.2f'), APoint(304.00 + dx_pillow, 254.35), 10).rotation = 1.57
            # начальная толщина подушки
            mSp.AddText(format(round(new_mark_pillow_fact[0] - new_mark_ditch_fact[0], 2), '.2f'), APoint(304.00 + dx_pillow, 210.00), 10).rotation = 1.57
            # начальный ПК в строке "Пикеты"
            pk_start = new_pk_int[0] + '+' + new_pk_float[0]
            mSp.AddText(pk_start, APoint(280.00 + dx_pillow, 32), 10)
            p = 304.00
            for i in range(len(new_distance_pk) - 1):
                pk_dist = round(float(new_distance_pk[i + 1]) - float(new_distance_pk[i]), 2)
                # проставка рисок в строке "Растояния"
                dh_x = p + pk_dist
                mSp.AddLine(APoint(dh_x + dx_pillow, 70.00), APoint(dh_x + dx_pillow, 50.00))
                # подпись расстояния
                otstup_text = pk_dist - (pk_dist / 2 + 5)
                mSp.AddText(format(pk_dist, '.2f'), APoint(dh_x - otstup_text + dx_pillow, 61.30), 10).rotation = 1.57
                p += pk_dist
                # фактическая отметка дна траншеии
                mSp.AddText(format(float(new_mark_ditch_fact[i + 1]), '.2f'), APoint(dh_x + dx_pillow, 354.33), 10).rotation = 1.57
                # проектная отметка верха подушки
                mSp.AddText(format(float(new_mark_pillow[i + 1]), '.2f'), APoint(dh_x + dx_pillow, 304.35), 10).rotation = 1.57
                # фактические отметки верха подушки
                mSp.AddText(format(new_mark_pillow_fact[i + 1], '.2f'), APoint(dh_x + dx_pillow, 254.35), 10).rotation = 1.57
                # толщина слоя подушки
                mSp.AddText(format(round(new_mark_pillow_fact[i + 1] - new_mark_ditch_fact[i + 1], 2), '.2f'), APoint(dh_x + dx_pillow, 210.00), 10).rotation = 1.57

                if len(new_distance_pk)-2 == i:
                    l_prof = float(new_distance_pk[-1]) - float(new_distance_pk[0]) - 15
                    pk_end = new_pk_int[-1] + '+' + new_pk_float[-1]
                    mSp.AddText(pk_end, APoint(304.00 + l_prof + dx_pillow, 32), 10)
                else:
                    piket = new_pk_int[i + 1]
                    cutoff = len(piket) - 1
                    if new_pk_float[i + 1] == '00':
                        mSp.AddLine(APoint(dh_x + dx_pillow, 100.00), APoint(dh_x + dx_pillow, 50.00))
                        if piket[cutoff:] == '0' or piket[cutoff:] == '5':
                            mSp.AddText(piket, APoint(dh_x - 15 + dx_pillow, 32), 10)
                        else:
                            mSp.AddText(piket[cutoff:], APoint(dh_x - 3.24 + dx_pillow, 32), 10)
            if new_pk_float[0] == '00':
                mSp.AddLine(APoint(304 + dx_pillow, 100.00), APoint(304 + dx_pillow, 50.00))

        def basement_all_pipe():
            # начальня риска в строке "Растояния"
            mSp.AddLine(APoint(304.00 + dx_pipe, 50.00), APoint(304.00 + dx_pipe, 70.00))
            # начальная отметка земли
            mSp.AddText(format(new_mark_earth[0], '.2f'), APoint(304.00 + dx_pipe, 404.33), 10).rotation = 1.57
            # начальная проектная отметка трубы
            mSp.AddText(format(new_mark_pipe[0], '.2f'), APoint(304.00 + dx_pipe, 354.35), 10).rotation = 1.57
            # начальная фактическая отметка трубы
            mSp.AddText(format(round(new_mark_pipe_fact[0], 2), '.2f'), APoint(304.00 + dx_pipe, 303.42), 10).rotation = 1.57
            # начальное отклонение (-0,03)
            mSp.AddText(format(new_mark_pipe_fact[0]-new_mark_pipe[0], '.2f'), APoint(304.00 + dx_pipe, 255.35), 10).rotation = 1.57
            # начальный ПК в строке "Пикеты"
            pk_start = new_pk_int[0] + '+' + new_pk_float[0]
            mSp.AddText(pk_start, APoint(280.00 + dx_pipe, 32), 10)
            p = 304.00
            for i in range(len(new_distance_pk) - 1):
                pk_dist = round(float(new_distance_pk[i + 1]) - float(new_distance_pk[i]), 2)
                # проставка рисок в строке "Растояния"
                dh_x = p + pk_dist
                mSp.AddLine(APoint(dh_x + dx_pipe, 70.00), APoint(dh_x + dx_pipe, 50.00))
                # подпись расстояния
                otstup_text = pk_dist - (pk_dist / 2 + 5)
                mSp.AddText(format(pk_dist, '.2f'), APoint(dh_x - otstup_text + dx_pipe, 61.30), 10).rotation = 1.57
                p += pk_dist
                # отметки верха земли
                mSp.AddText(format(float(new_mark_earth[i + 1]), '.2f'), APoint(dh_x + dx_pipe, 404.33), 10).rotation = 1.57
                # проектные отметки трубы
                mSp.AddText(format(float(new_mark_pipe[i + 1]), '.2f'), APoint(dh_x + dx_pipe, 354.35), 10).rotation = 1.57
                # фактические отметки трубы
                mSp.AddText(format(new_mark_pipe_fact[i + 1], '.2f'), APoint(dh_x + dx_pipe, 303.42), 10).rotation = 1.57
                # отклонения от проекта по трубе
                mSp.AddText(round(new_mark_pipe_fact[i + 1] - new_mark_pipe[i + 1], 2), APoint(dh_x + dx_pipe, 255.35), 10).rotation = 1.57

                if len(new_distance_pk)-2 == i:
                    l_prof = float(new_distance_pk[-1]) - float(new_distance_pk[0]) - 15
                    pk_end = new_pk_int[-1] + '+' + new_pk_float[-1]
                    mSp.AddText(pk_end, APoint(304.00 + l_prof + dx_pipe, 32), 10)
                else:
                    piket = new_pk_int[i + 1]
                    cutoff = len(piket) - 1
                    if new_pk_float[i + 1] == '00':
                        mSp.AddLine(APoint(dh_x + dx_pipe, 100.00), APoint(dh_x + dx_pipe, 50.00))
                        if piket[cutoff:] == '0' or piket[cutoff:] == '5':
                            mSp.AddText(piket, APoint(dh_x - 15 + dx_pipe, 32), 10)
                        else:
                            mSp.AddText(piket[cutoff:], APoint(dh_x - 3.24 + dx_pipe, 32), 10)
            if new_pk_float[0] == '00':
                mSp.AddLine(APoint(304 + dx_pipe, 100.00), APoint(304 + dx_pipe, 50.00))

        def basement_all_filling():

            # начальня риска в строке "Растояния"
            mSp.AddLine(APoint(304.00 + dx_filling, 50.00), APoint(304.00 + dx_filling, 70.00))
            # начальная проектная отметка верха обсыпки
            mSp.AddText(format(new_mark_filling[0], '.2f'), APoint(304.00 + dx_filling, 354.33), 10).rotation = 1.57
            # начальная фактическая отметка верха обсыпки
            mSp.AddText(format(new_mark_filling_fact[0], '.2f'), APoint(304.00 + dx_filling, 303.86), 10).rotation = 1.57
            # начальная фактическая отметка верха трубы
            mSp.AddText(format(round(new_mark_pipe_fact[0], 2), '.2f'), APoint(304.00 + dx_filling, 253.88), 10).rotation = 1.57
            # начальная толщина обсыпки
            mSp.AddText(format(round(new_mark_filling_fact[0] - new_mark_pipe_fact[0], 2), '.2f'), APoint(304.00 + dx_filling, 210.80), 10).rotation = 1.57
            # начальный ПК в строке "Пикеты"
            pk_start = new_pk_int[0] + '+' + new_pk_float[0]
            mSp.AddText(pk_start, APoint(280.00 + dx_filling, 32), 10)
            p = 304.00
            for i in range(len(new_distance_pk) - 1):
                pk_dist = round(float(new_distance_pk[i + 1]) - float(new_distance_pk[i]), 2)
                # проставка рисок в строке "Растояния"
                dh_x = p + pk_dist
                mSp.AddLine(APoint(dh_x + dx_filling, 70.00), APoint(dh_x + dx_filling, 50.00))
                # подпись расстояния
                otstup_text = pk_dist - (pk_dist / 2 + 5)
                mSp.AddText(format(pk_dist, '.2f'), APoint(dh_x - otstup_text + dx_filling, 61.30), 10).rotation = 1.57
                p += pk_dist
                # проектная отметка верха обсыпки
                mSp.AddText(format(float(new_mark_filling[i + 1]), '.2f'), APoint(dh_x + dx_filling, 354.33), 10).rotation = 1.57
                # фактическая отметка верха обсыпки
                mSp.AddText(format(float(new_mark_filling_fact[i + 1]), '.2f'), APoint(dh_x + dx_filling, 303.86), 10).rotation = 1.57
                # фактические отметки верха трубы
                mSp.AddText(format(new_mark_pipe_fact[i + 1], '.2f'), APoint(dh_x + dx_filling, 253.88), 10).rotation = 1.57
                # толщина слоя обсыпки
                mSp.AddText(format(round(new_mark_filling_fact[i + 1] - new_mark_pipe_fact[i + 1], 2), '.2f'),
                            APoint(dh_x + dx_filling, 210.80), 10).rotation = 1.57

                if len(new_distance_pk)-2 == i:
                    l_prof = float(new_distance_pk[-1]) - float(new_distance_pk[0]) - 15
                    pk_end = new_pk_int[-1] + '+' + new_pk_float[-1]
                    mSp.AddText(pk_end, APoint(304.00 + l_prof + dx_filling, 32), 10)
                else:
                    piket = new_pk_int[i + 1]
                    cutoff = len(piket) - 1
                    if new_pk_float[i + 1] == '00':
                        mSp.AddLine(APoint(dh_x + dx_filling, 100.00), APoint(dh_x + dx_filling, 50.00))
                        if piket[cutoff:] == '0' or piket[cutoff:] == '5':
                            mSp.AddText(piket, APoint(dh_x - 15 + dx_filling, 32), 10)
                        else:
                            mSp.AddText(piket[cutoff:], APoint(dh_x - 3.24 + dx_filling, 32), 10)
            if new_pk_float[0] == '00':
                mSp.AddLine(APoint(304 + dx_filling, 100.00), APoint(304 + dx_filling, 50.00))

        def height_scale(dx, r1, data_1, data_2):
            min_mark_pipe = int(round(min(data_1) - 2))
            max_mark_pipe = int(round(max(data_2) + 2))
            r = r1 + 20
            h = r1 + 25
            for q in range(max_mark_pipe - min_mark_pipe):
                mSp.AddText(min_mark_pipe, APoint((264 - 33) + dx, r), 10)
                mSp.AddLine(APoint(259 + dx, h), APoint(264 + dx, h))
                min_mark_pipe += 1
                r += 25
                h += 25
            mSp.AddLine(APoint(264 + dx, r1), APoint(264 + dx, r - 20))
            mSp.AddLine(APoint(264 + dx-5, r1), APoint(264 + dx-5, r - 20))

        def profile_line(g, dx, list_1, list_2, dx_transh_earth):
            min_mark = int(round(min(list_1)) - 2)
            x_1 = 264 + 40 + dx
            y_1 = g + (list_1[0] - min_mark) * 25 + 25

            for lp in range(len(list_1) - 1):
                x_2 = x_1 + (float(new_distance_pk[lp + 1]) - float(new_distance_pk[lp]))
                y_2 = y_1 + ((list_1[lp + 1] - list_1[lp]) * 25)
                mSp.AddLine(APoint(x_1, y_1), APoint(x_2, y_2))
                mSp.AddLine(APoint(x_1, g), APoint(x_1, y_1))
                x_1 = x_2
                y_1 = y_2
            mSp.AddLine(APoint(x_1, g), APoint(x_1, y_1, ))

            x_1 = 264 + 40 + dx
            y_1 = g + (list_1[0] - min_mark) * 25 + 25 + dx_transh_earth
            for lp in range(len(list_2) - 1):
                x_2 = x_1 + (float(new_distance_pk[lp + 1]) - float(new_distance_pk[lp]))
                y_2 = y_1 + ((list_2[lp + 1] - list_2[lp]) * 25)
                mSp.AddLine(APoint(x_1, y_1), APoint(x_2, y_2))
                x_1 = x_2
                y_1 = y_2

        def ditch():
            # простроение подвала разработки
            line_basement(450, 264, 0)  # контур шапки подвала (разработка)
            line_basement_gorizont(h_step, 8, 0)  # горизонтальные черты в шапке подвала (разработка)
            text_basement_hat(text_basement_transh, coordinates_text_basement_transh, 0)  # подпись строк в шапке подвала (разработка)
            mSp.AddLine(APoint(41.37, 177.55), APoint(244.74, 177.55))  # дробная черта 'окос/ширина'
            line_longitudinal(9, 0)  # продольные линии подвара (разработка)
            basement_all_ditch()  # запонение строк подвала (разработка)
            height_scale(0, 450, new_mark_ditch, new_mark_earth)  # шкала (разработка)
            profile_line(450, 0, new_mark_ditch, new_mark_earth, dx_ditch_earth)

        def pillow():
            line_basement(400, 264, dx_pillow)
            line_basement_gorizont(h_step, 7, dx_pillow)
            text_basement_hat(text_basement_pillow, coordinates_text_basement_pillow, dx_pillow)
            mSp.AddLine(APoint(41.37 + dx_pillow, 177.55), APoint(244.74 + dx_pillow, 177.55))  # дробная черта 'окос/ширина'
            line_longitudinal(8, dx_pillow)
            basement_all_pillow()
            height_scale(dx_pillow, 400, new_mark_ditch, new_mark_pillow)
            dx_pillow_ditch = (new_mark_pillow_fact[0] - new_mark_ditch_fact[0]) * 25
            profile_line(400, dx_pillow, new_mark_ditch, new_mark_pillow, dx_pillow_ditch)

        def pipe():
            # простроение подвала укладки
            line_basement(450, 264, dx_pipe)  # контур шапки подвала (укладка)
            line_basement_gorizont(h_step, 8, dx_pipe)  # горизонтальные черты в шапке подвала (укладка)
            text_basement_hat(text_basement_pipe, coordinates_text_basement_pipe, dx_pipe)  # подпись строк в шапке подвала (укладка)
            mSp.AddLine(APoint(41.37 + dx_pipe, 227.55), APoint(244.74 + dx_pipe, 227.55))  # дробная черта 'окос/ширина'
            line_longitudinal(9, dx_pipe)  # продольные линии подвара (укладка)
            basement_all_pipe()
            height_scale(dx_pipe, 450, new_mark_pipe, new_mark_earth)  # шкала (разработка)
            dx_pipe_earth = (new_mark_earth[0] - new_mark_pipe[0]) * 25
            profile_line(450, dx_pipe, new_mark_pipe, new_mark_earth, dx_pipe_earth)
            mSp.AddLine(APoint(0 + dx_pipe, 150.00), APoint(264.00 + dx_pipe, 100.00))  # дробная черта 'лина/уклон'

        def filling():
            line_basement(400, 264, dx_filling)
            line_basement_gorizont(h_step, 7, dx_filling)
            text_basement_hat(text_basement_filling, coordinates_text_basement_filling, dx_filling)
            mSp.AddLine(APoint(41.37 + dx_filling, 177.55), APoint(244.74 + dx_filling, 177.55))  # дробная черта 'окос/ширина'
            line_longitudinal(8, dx_filling)
            basement_all_filling()
            height_scale(dx_filling, 400, new_mark_pipe, new_mark_filling)
            dx_filling_pipe = (new_mark_filling[0] - new_mark_pipe[0]) * 25
            profile_line(400, dx_filling, new_mark_pipe, new_mark_filling, dx_filling_pipe)

        if chk_state_ditch.get() == True:
            ditch()
            style_text()
        bar['value'] = 25
        window.update_idletasks()

        if chk_state_pillow.get() == True:
            pillow()
            style_text()
        bar['value'] = 50
        window.update_idletasks()
        if chk_state_pipe.get() == True:
            pipe()
            style_text()
        bar['value'] = 75
        window.update_idletasks()
        if chk_state_filling.get() == True:
            filling()
            style_text()
        bar['value'] = 100
        window.update_idletasks()

    def add_mark_in_excel():

        start_import = str(input_start_import.get())
        point = start_import.split('+')
        insert_point = []
        dist = []
        time_out = 2
        acad = Autocad()

        def add_mark(text):
            dict_mark = {}
            while len(dict_mark) == 0:
                try:
                    for mark in acad.get_selection(text=text):
                        coord = int(float(mark.InsertionPoint[0]) * 1000)
                        dict_mark[coord] = float(mark.TextString)
                except:
                    dict_mark.clear()

            sorted_tuple_mark = sorted(dict_mark.items())
            dict_for_mark = dict(sorted_tuple_mark)
            return dict_for_mark

        def add_dist():

            while len(insert_point) == 0:
                try:
                    for point in acad.get_selection(text='Выдели риски расстояний'):
                        insert_point.append(point.EndPoint[0])
                except:
                    insert_point.clear()

        try:
            time.sleep(time_out)
            dict_mark_pipe = add_mark(text='Выдели отметки по трубе')
            dict_mark_earth = add_mark(text='Выдели отметки по земле')
            add_dist()

        except:
            mb.showinfo(
                "Внимание!",
                "Произошел сбой. Повторите функцию еще раз!")

        s_insert_point = sorted(insert_point)

        list_pk_int = []
        list_pk_float = []

        list_pk_int.append(float(point[0]))
        list_pk_float.append(float(point[1]))

        def calc(a, b):
            return round(b - a, 2)

        for i in range(len(insert_point)-1):
            calc_dist = calc(s_insert_point[i], s_insert_point[i+1])
            if calc_dist != 0:
                dist.append(calc_dist)

        str_pk_int = list_pk_int[0]
        str_pk_float = list_pk_float[0]
        for i in range(len(dist)):
            t = str_pk_float + dist[i]
            str_pk_float += dist[i]

            if round(t, 2) >= 100:
                str_pk_int += 1
                str_pk_float = 0
                list_pk_int.append(str_pk_int)
                list_pk_float.append(0)
            else:
                list_pk_float.append(t)
                list_pk_int.append(str_pk_int)

        file_import = f'{int(list_pk_int[0])}+{list_pk_float[0]} - {int(list_pk_int[-1])}+{list_pk_float[-1]}.xlsx'
        wb = openpyxl.Workbook(file_import)
        wb.save(file_import)
        wb_1 = openpyxl.reader.excel.load_workbook(filename=file_import)
        sheet = wb_1['Sheet']
        i = 1
        for j, l in zip(dict_mark_pipe, dict_mark_earth):
            mark_pipe = sheet[f'C{i}']
            mark_pipe.value = dict_mark_pipe[j]
            mark_earth = sheet[f'D{i}']
            mark_earth.value = dict_mark_earth[l]
            i += 1

        for i in range(len(list_pk_int)):
            pk_int = sheet[f'A{i + 1}']
            pk_int.value = list_pk_int[i]
        for i in range(len(list_pk_float)):
            pk_float = sheet[f'B{i + 1}']
            pk_float.value = list_pk_float[i]
        wb_1.save(file_import)

        if len(dist) > 0:
            acad.prompt(text='Готовчинко')
            mb.showinfo(
                "Добавлен файл",
                f"Добавлен файл: ПК{int(list_pk_int[0])}+{list_pk_float[0]} - ПК{int(list_pk_int[-1])}+{list_pk_float[-1]} ")

    def resource_path(relative):
        if hasattr(sys, "_MEIPASS"):
            return os.path.join(sys._MEIPASS, relative)
        return os.path.join(relative)

    try:
        # список доступных шрифтов
        list_styles = []
        for i in mSp_1.TextStyles:
            list_styles.append(i.name)
    except KeyError:
        mb.showerror(
            'Ошибка',
            f'Не удалось определить коллекцию шрифтов')

    window = Tk()
    window.title("Ленивчик")
    window.geometry('460x165')

    path_ico = resource_path('kl.ico')
    window.resizable(width=False, height=False)
    window.iconbitmap(path_ico)

    tab_control = ttk.Notebook(window)
    tab1 = ttk.Frame(tab_control)
    tab2 = ttk.Frame(tab_control)
    tab3 = ttk.Frame(tab_control)
    tab_control.add(tab1, text='Чертилка')
    tab_control.add(tab2, text='Расчет факта')
    tab_control.add(tab3, text='Импорт отметок')

    # список доступных шрифтов
    list_styles_font = Label(tab1, text="Шрифт:               ", font=('Arial Bold', 9))
    list_styles_font.grid(column=0, row=6)
    combo = Combobox(tab1, width=10)  # список доступных шрифтов
    combo['values'] = list_styles
    combo.current(0)  # установите вариант по умолчанию
    combo.grid(column=1, row=6)

    start_pk = Label(tab1, text="Начальный ПК:", font=('Arial Bold', 9))
    start_pk.grid(column=0, row=2)

    input_start_pk = Entry(tab1, width=13, text="")
    input_start_pk.grid(column=1, row=2)

    end_pk = Label(tab1, text="Конечный ПК:  ", font=('Arial Bold', 9))
    end_pk.grid(column=0, row=3)

    input_end_pk = Entry(tab1, width=13, text="")
    input_end_pk.grid(column=1, row=3)

    btn = Button(tab1, text="  Облениться совсем ", width=20, command=calc)
    btn.grid(column=2, row=7)

    end_pk = Label(tab1, text="Запусти AutoCAD", font=('Arial Bold', 9))
    end_pk.grid(column=2, row=6)

    btn_1 = Button(tab1, text="Открыть", command=file)
    btn_1.grid(column=3, row=0)

    lbl_1 = Label(tab1, text='Выбери файл для чертилки -->', font=('Arial Bold', 9), background="#ff9d00")
    lbl_1.grid(column=2, row=0)

    btn_1 = Button(tab2, text="Открыть", command=calc_file)
    btn_1.grid(column=3, row=0)

    lbl_5 = Label(tab2, text='Выбери файл для расчета -->', font=('Arial Bold', 9), background="#ff9d00")
    lbl_5.grid(column=2, row=0)

    chk_state_ditch = IntVar()
    chk_state_ditch.set(0)
    chk_ditch = Checkbutton(tab1, text='Разработка', var=chk_state_ditch)
    chk_ditch.grid(column=2, row=2)

    chk_state_pillow = IntVar()
    chk_state_pillow.set(0)
    chk_pillow = Checkbutton(tab1, text='Подушка', var=chk_state_pillow)
    chk_pillow.grid(column=3, row=2)

    chk_state_pipe = IntVar()
    chk_state_pipe.set(0)
    chk_pipe = Checkbutton(tab1, text='Укладка      ', var=chk_state_pipe)
    chk_pipe.grid(column=2, row=3)

    chk_state_filling = IntVar()
    chk_state_filling.set(0)
    chk = Checkbutton(tab1, text='Обсыпка', var=chk_state_filling)
    chk.grid(column=3, row=3)

    style = ttk.Style()
    bar = Progressbar(tab1, length=150, style='black.Horizontal.TProgressbar', mode='determinate', value=0)
    bar.grid(column=2, row=8)

    tab_control.pack(expand=1, fill='both')

    btn_1_3 = Button(tab3, text="Импорт", width=12, command=add_mark_in_excel)
    btn_1_3.grid(column=3, row=3)
    text_info = 'Выбери данные в последовательности:\n1 - отметки по трубе\n2 - отметки по земле' \
                '\n3 - отрезки по расстоянию'
    lbl_info = Label(tab3, text=text_info, font=('Arial Bold', 10))
    lbl_info.grid(column=3, row=4)

    lbl_start = Label(tab3, text="Стартовый ПК:", width=15, font=('Arial Bold', 10))
    lbl_start.grid(column=2, row=2)
    input_start_import = Entry(tab3, width=15, text="")
    input_start_import.grid(column=3, row=2)

    window.mainloop()
